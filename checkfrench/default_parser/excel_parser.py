"""
File        : excel_parser.py
Author      : Silous
Created on  : 2025-07-19
Description : Parser for excel files.

This module provides a function to parse a excel file and return its non-empty lines.
This parser acts as a default parser for excel files.
"""


# == Imports ==================================================================

from logging import Logger

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

from checkfrench.logger import get_logger

# == Global Variables =========================================================

logger: Logger = get_logger(__name__)


# == Functions ================================================================

def parse_file(pathfile: str, argument: str) -> list[tuple[str, str]]:
    """Parse an Excel file and return each non-empty cell from the specified column with row identifier.

    Args:
        pathfile (str): Path to the Excel file (.xlsx).
        argument (str): Column(s), e.g., 'E' or 'E,A'. First is the content column,
                        second (optional) is the row ID column.

    Returns:
        list[tuple[str, str]]: List of (row ID as string, cell content).
    """
    try:
        parts: list[str] = [arg.strip().upper() for arg in argument.split(",")]
        col_value_index: int = column_index_from_string(parts[0])  # Always required
        col_id_index: int | None = column_index_from_string(parts[1]) if len(parts) > 1 else None
    except ValueError:
        logger.error("%s is not a valid argument for the excel parser.", argument)
        return []

    try:
        wb: Workbook = load_workbook(pathfile, data_only=True)
        ws = wb.active

        results: list[tuple[str, str]] = []

        if ws is not None:
            for row in ws.iter_rows(min_row=1):
                if col_value_index > len(row):
                    continue  # Value column out of range

                cell_value = row[col_value_index - 1]

                if cell_value.value is not None and str(cell_value.value).strip():
                    if col_id_index is not None and col_id_index <= len(row):
                        cell_id = row[col_id_index - 1]
                        row_id: str = str(cell_id.value).strip() if cell_id.value is not None else str(cell_value.row)
                    else:
                        row_id = str(cell_value.row)

                    results.append((row_id, str(cell_value.value)))

        return results

    except Exception as e:
        logger.error("Error when parsing the Excel file %s : %s", pathfile, e)
        return []
