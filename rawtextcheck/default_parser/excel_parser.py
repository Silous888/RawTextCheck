"""
File        : excel_parser.py
Author      : Silous
Created on  : 2025-07-19
Description : Parser for excel files.

This module provides a function to parse a excel file and return non-empty cells of a column.
This parser acts as a default parser for excel files.
"""


# == Imports ==================================================================

from logging import Logger

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

from rawtextcheck.logger import get_logger
from rawtextcheck.newtype import ParserArgument


# == Constants ================================================================

COL_ARG = ParserArgument(name="col", optional=False)
COL_ID_ARG = ParserArgument(name="colID", optional=True)

LIST_ARGUMENTS: list[ParserArgument] = [COL_ARG, COL_ID_ARG]


# == Global Variables =========================================================

logger: Logger = get_logger(__name__)


# == Functions ================================================================

def parse_file(filepath: str, arguments: dict[str, str]) -> list[tuple[str, str]]:
    """Parse an Excel file and return each non-empty cell from the specified column with row identifier.

    Args:
        filepath (str): Path to the Excel file (.xlsx).
        argument (dict[str, str]): Specific argument for this file.
            keys:
                - "col": Column letter (e.g., "A") to parse.
                - "colID": Optional column letter for row identifier (default is the row number).

    Returns:
        list[tuple[str, str]]: List of (row ID as string, cell content).
    """
    try:
        col_value_index: int = column_index_from_string(arguments[COL_ARG.name])
        col_id_index: int | None = None
        if COL_ID_ARG.name in arguments.keys():
            col_id_index = column_index_from_string(arguments[COL_ID_ARG.name])
    except ValueError:
        logger.error("%s is not a valid argument for the excel parser.", arguments)
        return []

    try:
        wb: Workbook = load_workbook(filepath, data_only=True)
        ws = wb.active

        results: list[tuple[str, str]] = []

        if ws is not None:
            for row in ws.iter_rows(min_row=1):
                if col_value_index > len(row):
                    continue  # Value column out of range

                cell_value = row[col_value_index - 1]

                if cell_value.value is not None and str(cell_value.value).strip():
                    if col_id_index is not None and col_id_index <= len(row):
                        cell_id = row[col_id_index - 1]
                        row_id: str = str(cell_id.value).strip() if cell_id.value is not None else str(cell_value.row)
                    else:
                        row_id = str(cell_value.row)

                    results.append((row_id, str(cell_value.value)))

        return results

    except Exception as e:
        logger.error("Error when parsing the Excel file %s : %s", filepath, e)
        return []
